"""导出服务 — Excel/CSV 巡检报告生成与历史管理。"""
from __future__ import annotations

import csv
import io
import json
import os
import time
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..adapters import get_query_sets
from ..config import Settings
from ..core.constants import OS_CHECK_LABELS, QUERY_LABELS
from .cache import CacheStore, get_cache
from .server_service import ServerService, get_server_service

EXPORT_CATEGORIES = [
    {"id": "basic_info", "label": "基础信息", "icon": "bi-info-circle"},
    {"id": "db_info", "label": "数据库信息", "icon": "bi-database"},
    {"id": "storage", "label": "存储空间", "icon": "bi-hdd"},
    {"id": "objects", "label": "对象统计", "icon": "bi-grid"},
    {"id": "performance", "label": "性能监控", "icon": "bi-graph-up"},
    {"id": "install_path", "label": "安装路径", "icon": "bi-folder"},
    {"id": "db_log_errors", "label": "数据库日志", "icon": "bi-file-earmark-text"},
]

EXPORT_OS_CHECKS = [
    {"id": "memory", "label": "系统内存", "icon": "bi-memory"},
    {"id": "disk", "label": "系统磁盘", "icon": "bi-hdd-stack"},
    {"id": "cpu", "label": "CPU负载", "icon": "bi-cpu"},
    {"id": "os_errors", "label": "系统日志错误", "icon": "bi-exclamation-diamond"},
]


class ExportService:
    def __init__(self, settings: Settings, server_service: ServerService, cache: CacheStore):
        self.settings = settings
        self.server_service = server_service
        self.cache = cache

    # ═══════════════ CSV ═══════════════
    def export_csv(self, server_id: str) -> tuple[str, str]:
        server = self.server_service.get(server_id)
        if not server:
            raise ValueError("服务不存在")
        cached = self.cache.get(server_id)
        if not cached:
            raise ValueError("暂无数据，请先采集")

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["服务器", server.get("name") or server.get("ssh_host", "")])
        writer.writerow(["采集时间", cached.get("timestamp", "")])
        writer.writerow([])

        os_info = cached.get("os_info", {})
        if os_info:
            writer.writerow(["=== 操作系统信息 ==="])
            for ck, cr in os_info.items():
                writer.writerow([OS_CHECK_LABELS.get(ck, ck)])
                if cr.get("columns") and cr.get("rows"):
                    writer.writerow(cr["columns"])
                    for row in cr["rows"]:
                        writer.writerow(row)
                elif cr.get("output"):
                    writer.writerow([cr["output"]])
                writer.writerow([])

        db_queries = cached.get("db_queries", {})
        query_sets = get_query_sets(server.get("db_type"))
        for cat, queries in db_queries.items():
            writer.writerow([f'=== {query_sets.get(cat, {}).get("label", cat)} ==='])
            for qn, qr in queries.items():
                writer.writerow([QUERY_LABELS.get(qn, qn)])
                if qr.get("columns") and qr.get("rows"):
                    writer.writerow(qr["columns"])
                    for row in qr["rows"]:
                        writer.writerow(row)
                writer.writerow([])

        filename = f"{server_id}_export.csv"
        return filename, output.getvalue()

    # ═══════════════ 巡检 Excel ═══════════════
    def export_xlsx(self, server_ids: list[str], categories: list[str], os_checks: list[str], organize_by: str = "server") -> str:
        if not server_ids:
            raise ValueError("请选择至少一台服务器")
        if not categories and not os_checks:
            raise ValueError("请选择至少一项导出内容")

        servers = [s for s in self.server_service.list() if s.get("id") in server_ids]
        if not servers:
            raise ValueError("未找到所选服务器")

        wb = Workbook()
        wb.remove(wb.active)

        header_font = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
        section_font = Font(name="Microsoft YaHei", bold=True, size=12, color="4F46E5")
        thin_border = Border(
            left=Side(style="thin", color="D4D4D8"),
            right=Side(style="thin", color="D4D4D8"),
            top=Side(style="thin", color="D4D4D8"),
            bottom=Side(style="thin", color="D4D4D8"),
        )
        wrap_align = Alignment(wrap_text=True, vertical="top")

        if organize_by == "server":
            for server in servers:
                self._fill_server_sheet(wb, server, categories, os_checks, header_font, header_fill, section_font, thin_border, wrap_align)
            self._fill_summary_sheet(wb, servers, header_font, header_fill, section_font, thin_border)
        else:
            for cat in categories:
                qs = get_query_sets(servers[0].get("db_type"))
                if cat in qs:
                    ws = wb.create_sheet(title=qs[cat]["label"][:31])
                    self._fill_category_sheet(ws, servers, cat, "db", header_font, header_fill, section_font, thin_border, wrap_align)
            for chk in os_checks:
                if chk in OS_CHECK_LABELS:
                    ws = wb.create_sheet(title=OS_CHECK_LABELS[chk][:31])
                    self._fill_category_sheet(ws, servers, chk, "os", header_font, header_fill, section_font, thin_border, wrap_align)

        self.settings.reports_dir.mkdir(parents=True, exist_ok=True)
        filename = f"巡检报告_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = self.settings.reports_dir / filename
        wb.save(filepath)
        self._add_export_history(filename, len(servers), len(categories) + len(os_checks))
        return filename

    # ═══════════════ 内部填充 ═══════════════
    def _fill_server_sheet(self, wb, server, categories, os_checks, hf, hfill, sf, border, align):
        name = server.get("name") or server.get("ssh_host", "")
        sheet_name = (name[:28] + "..") if len(name) > 31 else name
        ws = wb.create_sheet(title=sheet_name[:31])

        cached = self.cache.get(server.get("id"))
        row = 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        cell = ws.cell(row=row, column=1, value=f"📊 巡检报告 — {name}")
        cell.font = Font(name="Microsoft YaHei", bold=True, size=14, color="4F46E5")
        row += 1
        ts = cached.get("timestamp", "暂无数据") if cached else "暂无数据"
        ws.cell(row=row, column=1, value=f"采集时间: {ts}").font = Font(color="64748B", size=10)
        row += 2

        if not cached:
            ws.cell(row=row, column=1, value="暂无采集数据，请先执行采集").font = Font(color="DC2626", size=11)
            self._auto_width(ws)
            return

        os_info = cached.get("os_info", {})
        db_queries = cached.get("db_queries", {})
        qs = get_query_sets(server.get("db_type"))

        for chk in os_checks:
            if chk in os_info and chk in OS_CHECK_LABELS:
                row = self._write_section_header(ws, row, f"🖥 {OS_CHECK_LABELS[chk]}", sf)
                row = self._write_check_result(ws, row, os_info[chk], hf, hfill, border, align)
                row += 1

        for cat in categories:
            if cat in db_queries and cat in qs:
                row = self._write_section_header(ws, row, f"📋 {qs[cat]['label']}", sf)
                for qname, qresult in db_queries[cat].items():
                    label = QUERY_LABELS.get(qname, qname)
                    row = self._write_section_header(ws, row, label, Font(name="Microsoft YaHei", bold=True, size=10, color="6366F1"))
                    row = self._write_query_result(ws, row, qresult, hf, hfill, border, align)
                    row += 1
            elif cat in os_info and cat in OS_CHECK_LABELS:
                row = self._write_section_header(ws, row, f"📋 {OS_CHECK_LABELS[cat]}", sf)
                row = self._write_check_result(ws, row, os_info[cat], hf, hfill, border, align)
                row += 1

        self._auto_width(ws)

    def _fill_category_sheet(self, ws, servers, key, stype, hf, hfill, sf, border, align):
        row = 1
        qs = get_query_sets(servers[0].get("db_type")) if servers else {}
        label = qs[key]["label"] if stype == "db" else OS_CHECK_LABELS.get(key, key)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws.cell(row=row, column=1, value=f"📊 {label}").font = Font(name="Microsoft YaHei", bold=True, size=14, color="4F46E5")
        row += 2

        for server in servers:
            name = server.get("name") or server.get("ssh_host", "")
            cached = self.cache.get(server.get("id"))
            row = self._write_section_header(ws, row, f"🖥 {name}", sf)
            if not cached:
                ws.cell(row=row, column=1, value="暂无数据").font = Font(color="DC2626", size=10)
                row += 2
                continue

            if stype == "os":
                chk_data = cached.get("os_info", {}).get(key)
                if chk_data:
                    row = self._write_check_result(ws, row, chk_data, hf, hfill, border, align)
                else:
                    ws.cell(row=row, column=1, value="(未采集)").font = Font(color="94A3B8", size=10)
                    row += 1
            else:
                if key in cached.get("db_queries", {}):
                    for qname, qresult in cached["db_queries"][key].items():
                        qlabel = QUERY_LABELS.get(qname, qname)
                        ws.cell(row=row, column=1, value=qlabel).font = Font(name="Microsoft YaHei", bold=True, size=10, color="6366F1")
                        row += 1
                        row = self._write_query_result(ws, row, qresult, hf, hfill, border, align)
                elif key in cached.get("os_info", {}):
                    row = self._write_check_result(ws, row, cached["os_info"][key], hf, hfill, border, align)
                else:
                    ws.cell(row=row, column=1, value="(未采集)").font = Font(color="94A3B8", size=10)
                    row += 1
            row += 1
        self._auto_width(ws)

    def _fill_summary_sheet(self, wb, servers, hf, hfill, sf, border):
        ws = wb.create_sheet(title="汇总概览")
        row = 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value="📊 巡检汇总报告").font = Font(name="Microsoft YaHei", bold=True, size=14, color="4F46E5")
        row += 1
        ws.cell(row=row, column=1, value=f"导出时间: {time.strftime('%Y-%m-%d %H:%M:%S')}").font = Font(color="64748B", size=10)
        row += 2

        headers = ["服务器", "采集时间", "在线状态", "连接数", "死锁数", "慢SQL数"]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=ci, value=h)
            cell.font = hf
            cell.fill = hfill
            cell.border = border
        row += 1

        for server in servers:
            name = server.get("name") or server.get("ssh_host", "")
            cached = self.cache.get(server.get("id"))
            ts = cached.get("timestamp", "-") if cached else "-"
            sessions = deadlocks = slow_count = "-"
            if cached:
                perf = cached.get("db_queries", {}).get("performance", {})
                sc = perf.get("session_count", {}).get("rows")
                if sc and sc[0]:
                    sessions = str(sc[0][0])
                dl = perf.get("deadlock_count", {}).get("rows")
                if dl and dl[0]:
                    deadlocks = str(dl[0][0])
                sl = perf.get("slow_sql", {}).get("rows", [])
                slow_count = str(len(sl))

            vals = [name, ts, "在线" if cached else "未采集", sessions, deadlocks, slow_count]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=ci, value=v)
                cell.border = border
                cell.alignment = Alignment(vertical="top")
            row += 1
        self._auto_width(ws)

    # ═══════════════ 样式辅助 ═══════════════
    def _write_section_header(self, ws, row, text, font):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws.cell(row=row, column=1, value=text).font = font
        return row + 1

    def _write_query_result(self, ws, row, qr, hf, hfill, border, align):
        if not qr or not qr.get("columns"):
            ws.cell(row=row, column=1, value="(无数据)").font = Font(color="94A3B8", size=10)
            return row + 1
        for ci, col in enumerate(qr["columns"], 1):
            cell = ws.cell(row=row, column=ci, value=str(col))
            cell.font = hf
            cell.fill = hfill
            cell.border = border
        row += 1
        for r_data in qr.get("rows") or []:
            for ci, val in enumerate(r_data, 1):
                cell = ws.cell(row=row, column=ci, value=str(val) if val is not None else "")
                cell.border = border
                cell.alignment = align
            row += 1
        return row

    def _write_check_result(self, ws, row, cr, hf, hfill, border, align):
        if cr.get("columns") and cr.get("rows"):
            return self._write_query_result(ws, row, cr, hf, hfill, border, align)
        output = cr.get("output", "")
        if output:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            cell = ws.cell(row=row, column=1, value=output[:3000])
            cell.font = Font(name="Consolas", size=9, color="334155")
            cell.alignment = Alignment(wrap_text=True)
            return row + 1
        ws.cell(row=row, column=1, value="(无数据)").font = Font(color="94A3B8", size=10)
        return row + 1

    def _auto_width(self, ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    # ═══════════════ 历史 ═══════════════
    def _add_export_history(self, filename, server_count, item_count):
        history_file = self.settings.reports_dir / "history.json"
        history: list = []
        if history_file.exists():
            try:
                with open(history_file, "r", encoding="utf-8") as f:
                    history = json.load(f)
            except (json.JSONDecodeError, OSError):
                history = []
        history.insert(0, {
            "filename": filename,
            "server_count": server_count,
            "item_count": item_count,
            "time": time.strftime("%Y-%m-%d %H:%M:%S"),
            "size": os.path.getsize(self.settings.reports_dir / filename),
        })
        if len(history) > 100:
            old = history.pop()
            old_path = self.settings.reports_dir / old["filename"]
            if old_path.exists():
                old_path.unlink()
        with open(history_file, "w", encoding="utf-8") as f:
            json.dump(history, f, ensure_ascii=False, indent=2)

    def history(self) -> list[dict[str, Any]]:
        history_file = self.settings.reports_dir / "history.json"
        if not history_file.exists():
            return []
        with open(history_file, "r", encoding="utf-8") as f:
            history = json.load(f)
        for h in history:
            h["size_str"] = self._format_size(h.get("size", 0))
        return history

    def delete_history(self, filename: str) -> None:
        """删除历史导出文件；用 basename + resolve 校验防止路径穿越。"""
        import os

        base = self.settings.reports_dir.resolve()
        safe_name = os.path.basename(filename)
        filepath = (base / safe_name).resolve()
        if base not in filepath.parents:
            return
        if filepath.is_file():
            filepath.unlink()
        history_file = base / "history.json"
        if history_file.exists():
            with open(history_file, "r", encoding="utf-8") as f:
                history = json.load(f)
            history = [h for h in history if h.get("filename") != safe_name]
            with open(history_file, "w", encoding="utf-8") as f:
                json.dump(history, f, ensure_ascii=False, indent=2)

    @staticmethod
    def _format_size(num: int) -> str:
        for unit in ("B", "KB", "MB", "GB"):
            if num < 1024:
                return f"{num:.1f} {unit}"
            num /= 1024
        return f"{num:.1f} TB"


_export_service: ExportService | None = None


def get_export_service(settings: Settings) -> ExportService:
    global _export_service
    if _export_service is None:
        _export_service = ExportService(settings, get_server_service(settings), get_cache())
    return _export_service
