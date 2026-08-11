import os
import sys
import json
import uuid
import re
import threading
import time
import traceback
from datetime import datetime
from functools import wraps
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import Flask, render_template, request, jsonify, redirect, url_for, session, Response, send_file

from auth import check_login, change_password, add_user, update_user, delete_user, load_users as load_auth_users, has_permission, PERMISSIONS

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

DATA_DIR = os.path.join(BASE_DIR, 'data')
if getattr(sys, 'frozen', False):
    DATA_DIR = os.path.join(os.path.dirname(sys.executable), 'data')
    BASE_DIR = sys._MEIPASS

app = Flask(__name__,
            template_folder=os.path.join(BASE_DIR, 'templates'),
            static_folder=os.path.join(BASE_DIR, 'static'))
app.config['SECRET_KEY'] = os.urandom(24).hex()
app.config['JSON_AS_ASCII'] = False
app.config['PERMANENT_SESSION_LIFETIME'] = 1800
app.jinja_env.auto_reload = True


@app.route('/favicon.ico')
@app.route('/favicon.svg')
def favicon():
    return send_file(os.path.join(BASE_DIR, 'static', 'favicon.svg'), mimetype='image/svg+xml')


scheduler = None
_collect_executor = None
_persist_executor = None


def _get_executor():
    global _collect_executor
    if _collect_executor is None:
        cfg = load_config()
        workers = max(1, min(20, cfg.get('collect_workers', 8)))
        _collect_executor = ThreadPoolExecutor(max_workers=workers)
    return _collect_executor


def _get_persist_executor():
    global _persist_executor
    if _persist_executor is None:
        _persist_executor = ThreadPoolExecutor(max_workers=4, thread_name_prefix='persist')
    return _persist_executor


def _reinit_executor():
    global _collect_executor, _persist_executor
    if _collect_executor:
        _collect_executor.shutdown(wait=False)
    _collect_executor = None
    if _persist_executor:
        _persist_executor.shutdown(wait=False)
    _persist_executor = None


try:
    from apscheduler.schedulers.background import BackgroundScheduler
    scheduler = BackgroundScheduler()
    scheduler.start()
except Exception:
    pass

from collector import collect_all, test_connection, db_control, app_control, QUERY_SETS, QUERY_LABELS, OS_CHECKS, OS_CHECK_LABELS
from collector import _ssh_connect, _ssh_exec, _need_ssh, translate_error, SSH_ERROR_TRANSLATE, DB_ERROR_TRANSLATE, SSH_FIX_WIN, SSH_FIX_LINUX
from persist import load_config, save_config, persist_os_error, persist_db_error, persist_slow_sql, cleanup_old_logs, query_logs
from db_config import load_servers_from_db, save_server_to_db, delete_server_from_db
CONFIG_FILE = os.path.join(DATA_DIR, 'servers.json')
CONFIG_LOCK = threading.Lock()

CACHE = {}
CACHE_LOCK = threading.Lock()

# 趋势历史：每个服务器保留最近 288 条 (24h@5min)
TREND_HISTORY = {}
TREND_LOCK = threading.Lock()
MAX_TREND_POINTS = 288

# 登录限速：{ip: [timestamp, ...]}
LOGIN_ATTEMPTS = {}
LOCKED_IPS = {}  # {ip: {'since': ts, 'username': str, 'count': int}}
LOGIN_LOCK = threading.Lock()
LOGIN_MAX_FAILURES = 5
LOGIN_WINDOW = 300       # 5 分钟窗口
LOGIN_BAN_DURATION = 900  # 封禁 15 分钟


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            if request.path.startswith('/api/'):
                return jsonify({"error": "未登录或会话已过期", "redirect": url_for('login')}), 401
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def permission_required(perm):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            user = session.get('user', {})
            if not has_permission(user, perm):
                if request.path.startswith('/api/'):
                    return jsonify({"error": "无权限"}), 403
                return redirect(url_for('index'))
            return f(*args, **kwargs)
        return decorated
    return decorator


def any_permission_required(*perms):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            user = session.get('user', {})
            if user.get('is_admin'):
                return f(*args, **kwargs)
            if not any(p in user.get('perms', []) for p in perms):
                if request.path.startswith('/api/'):
                    return jsonify({"error": "无权限"}), 403
                return redirect(url_for('index'))
            return f(*args, **kwargs)
        return decorated
    return decorator


def load_servers():
    db_servers = load_servers_from_db()
    if db_servers is not None and len(db_servers) > 0:
        return db_servers
    with CONFIG_LOCK:
        if not os.path.exists(CONFIG_FILE):
            return []
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            servers = json.load(f)
    for s in servers:
        if 'enabled_categories' not in s:
            s['enabled_categories'] = list(QUERY_SETS.keys())
        if 'enabled_os_checks' not in s:
            s['enabled_os_checks'] = list(OS_CHECKS.keys())
        if 'in_control' not in s:
            s['in_control'] = True
        if 'apps' not in s:
            s['apps'] = []
        if 'persist_enabled' not in s:
            s['persist_enabled'] = False
    if db_servers is not None and len(db_servers) == 0 and len(servers) > 0:
        save_servers(servers)
    return servers


def save_servers(servers):
    os.makedirs(os.path.dirname(CONFIG_FILE), exist_ok=True)
    with CONFIG_LOCK:
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(servers, f, ensure_ascii=False, indent=2)
    for s in servers:
        save_server_to_db(s)


def get_server_by_id(server_id):
    servers = load_servers()
    for s in servers:
        if s.get('id') == server_id:
            return s
    return None


def template_context():
    user = session.get('user', {})
    return {
        "username": session.get('username'),
        "is_admin": user.get('is_admin', False),
        "user_perms": user.get('perms', []),
    }


@app.before_request
def refresh_session_user():
    if session.get('logged_in') and session.get('username'):
        if request.path == '/api/stream':
            return
        now = time.time()
        last = session.get('last_activity', now)
        if now - last > 1800:
            session.clear()
            if request.path.startswith('/api/'):
                return jsonify({"error": "会话已过期", "redirect": url_for('login')}), 401
            return redirect(url_for('login'))
        session['last_activity'] = now
        # 仅每 30 秒刷新一次用户权限（auth.load_users 内部有 60 秒缓存兜底）
        if now - session.get('_user_refresh', 0) > 30:
            try:
                from auth import load_users as reload_users
                current = next((u for u in reload_users() if u.get('username') == session.get('username')), None)
                if current:
                    session['user'] = current
                    session['_user_refresh'] = now
            except Exception:
                pass


@app.errorhandler(500)
def internal_error(e):
    import logging
    logging.getLogger('oscar_monitor').error(f"500 Internal Server Error: {traceback.format_exc()}")
    return "<h2>500 内部服务器错误</h2><p>服务器遇到错误，请稍后重试或联系管理员。</p>", 500


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        client_ip = request.remote_addr or '127.0.0.1'
        username = request.form.get('username', '')
        password = request.form.get('password', '')

        # 先尝试验证（admin 不受限速影响）
        user = check_login(username, password)

        # ── 登录限速检查（admin 用户豁免）──
        is_admin_user = user and user.get('is_admin')
        if not is_admin_user:
            with LOGIN_LOCK:
                now_ts = time.time()
                attempts = [t for t in LOGIN_ATTEMPTS.get(client_ip, []) if now_ts - t < LOGIN_WINDOW]
                if len(attempts) >= LOGIN_MAX_FAILURES:
                    oldest = min(attempts) if attempts else 0
                    if now_ts - oldest < LOGIN_BAN_DURATION:
                        remaining = int((LOGIN_BAN_DURATION - (now_ts - oldest)) / 60) + 1
                        LOCKED_IPS[client_ip] = {'since': oldest, 'username': username or '?', 'count': len(attempts)}
                        session['login_error'] = f'登录失败次数过多，请 {remaining} 分钟后再试'
                        return redirect(url_for('login'))
                    else:
                        attempts = []
                        LOCKED_IPS.pop(client_ip, None)
                LOGIN_ATTEMPTS[client_ip] = attempts

        if user:
            with LOGIN_LOCK:
                LOGIN_ATTEMPTS.pop(client_ip, None)
                LOCKED_IPS.pop(client_ip, None)
            session['logged_in'] = True
            session['username'] = username
            session['user'] = user
            session['last_activity'] = time.time()
            session.permanent = True
            return redirect(url_for('index'))
        with LOGIN_LOCK:
            LOGIN_ATTEMPTS.setdefault(client_ip, []).append(time.time())
        session['login_error'] = '用户名或密码错误'
        return redirect(url_for('login'))

    # GET 请求：从 session 取出错误信息（避免刷新时重复提交表单）
    error = session.pop('login_error', None)
    return render_template('login.html', error=error)


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.route('/api/admin/locked-ips')
@login_required
@permission_required('admin')
def api_locked_ips():
    """返回当前被封锁的 IP 列表"""
    now_ts = time.time()
    result = []
    with LOGIN_LOCK:
        for ip, info in list(LOCKED_IPS.items()):
            since = info.get('since', 0)
            if now_ts - since < LOGIN_BAN_DURATION:
                remaining = max(0, int((LOGIN_BAN_DURATION - (now_ts - since)) / 60))
                result.append({
                    'ip': ip,
                    'username': info.get('username', '?'),
                    'count': info.get('count', 0),
                    'since': time.strftime('%H:%M:%S', time.localtime(since)),
                    'remaining_min': remaining
                })
            else:
                LOCKED_IPS.pop(ip, None)
    return jsonify({'locked': result})


@app.route('/api/admin/unlock-ip', methods=['POST'])
@login_required
@permission_required('admin')
def api_unlock_ip():
    """解锁指定 IP"""
    data = request.get_json(silent=True) or {}
    ip = data.get('ip', '').strip()
    if not ip:
        return jsonify({'ok': False, 'msg': 'IP 不能为空'})
    with LOGIN_LOCK:
        LOGIN_ATTEMPTS.pop(ip, None)
        LOCKED_IPS.pop(ip, None)
    return jsonify({'ok': True, 'msg': f'已解锁 {ip}'})


@app.route('/help')
@login_required
def help_page():
    return render_template('help.html', **template_context())


@app.route('/sql-terminal')
@login_required
@permission_required('admin')
def sql_terminal_page():
    servers = load_servers()
    return render_template('sql_terminal.html', servers=servers, **template_context())


@app.route('/')
@login_required
@permission_required('dashboard')
def index():
    try:
        servers = load_servers()
        return render_template('index.html', servers=servers, **template_context())
    except Exception as e:
        return f"<pre>首页渲染错误:\n{traceback.format_exc()}</pre>", 500


@app.route('/servers')
@login_required
@any_permission_required('servers_view', 'servers_edit')
def servers_page():
    try:
        servers = load_servers()
        return render_template('servers.html', servers=servers,
                               query_sets=QUERY_SETS, os_check_labels=OS_CHECK_LABELS,
                               log_enabled=load_config().get('log_enabled', False),
                               **template_context())
    except Exception as e:
        return f"<pre>服务管理页渲染错误:\n{traceback.format_exc()}</pre>", 500


@app.route('/servers/add')
@login_required
@permission_required('servers_edit')
def servers_add():
    try:
        return render_template('servers_add.html',
                               query_sets=QUERY_SETS, os_check_labels=OS_CHECK_LABELS,
                               log_enabled=load_config().get('log_enabled', False),
                               **template_context())
    except Exception as e:
        return f"<pre>添加服务器页渲染错误:\n{traceback.format_exc()}</pre>", 500


@app.route('/server/<server_id>')
@login_required
@any_permission_required('servers_view', 'servers_edit')
def server_detail(server_id):
    server = get_server_by_id(server_id)
    if not server:
        return redirect(url_for('servers_page'))
    try:
        return render_template('detail.html', server=server,
                               query_sets=QUERY_SETS, os_check_labels=OS_CHECK_LABELS,
                               query_labels=QUERY_LABELS, **template_context())
    except Exception as e:
        return f"<pre>详情页渲染错误:\n{traceback.format_exc()}</pre>", 500


@app.route('/control')
@login_required
@any_permission_required('control_view', 'control_exec')
def control_page():
    try:
        servers = load_servers()
        # 预计算应用分组
        app_groups = {}
        for s in servers:
            for app in s.get('apps', []):
                if app.get('in_control'):
                    g = app.get('group', '其他应用') or '其他应用'
                    if g not in app_groups:
                        app_groups[g] = []
                    app_groups[g].append({'server': s, 'app': app})
        return render_template('control.html', servers=servers, app_groups=app_groups, **template_context())
    except Exception as e:
        return f"<pre>管控页渲染错误:\n{traceback.format_exc()}</pre>", 500


@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    if request.method == 'POST':
        username = session.get('username')
        old_pwd = request.form.get('old_password', '')
        new_pwd = request.form.get('new_password', '')
        ok, msg = change_password(username, old_pwd, new_pwd)
        return render_template('profile.html', msg=msg, ok=ok, **template_context())
    return render_template('profile.html', msg=None, ok=None, **template_context())


@app.route('/users')
@login_required
@permission_required('admin')
def users_page():
    try:
        users = load_auth_users()
        return render_template('users.html', users=users, permissions=PERMISSIONS, **template_context())
    except Exception as e:
        return f"<pre>用户管理页渲染错误:\n{traceback.format_exc()}</pre>", 500


@app.route('/api/users', methods=['POST'])
@login_required
@permission_required('admin')
def api_add_user():
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "无效数据"}), 400
    ok, msg = add_user(data.get('username', ''), data.get('password', ''),
                       data.get('is_admin', False), data.get('perms', []))
    return jsonify({"status": "ok" if ok else "error", "msg": msg})


@app.route('/api/users/<username>', methods=['PUT'])
@login_required
@permission_required('admin')
def api_update_user(username):
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "无效数据"}), 400
    ok, msg = update_user(username, data)
    return jsonify({"status": "ok" if ok else "error", "msg": msg})


@app.route('/api/users/<username>', methods=['DELETE'])
@login_required
@permission_required('admin')
def api_delete_user(username):
    ok, msg = delete_user(username)
    return jsonify({"status": "ok" if ok else "error", "msg": msg})


@app.route('/api/test-connection', methods=['POST'])
@login_required
def api_test_connection():
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "无效的请求数据"}), 400
    try:
        result = test_connection(data)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500


@app.route('/api/servers', methods=['GET'])
@login_required
def api_list_servers():
    servers = load_servers()
    safe_servers = [{k: v for k, v in s.items() if k not in ('ssh_pass', 'db_pass')} for s in servers]
    return jsonify(safe_servers)


@app.route('/api/servers', methods=['POST'])
@login_required
@permission_required('servers_edit')
def api_add_server():
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "无效的请求数据"}), 400
    if data.get('persist_enabled') and not load_config().get('log_enabled'):
        return jsonify({"error": "全局日志持久化未启用，请先在系统配置中启用"}), 400
    servers = load_servers()
    data['id'] = uuid.uuid4().hex[:8]
    if 'enabled_categories' not in data:
        data['enabled_categories'] = list(QUERY_SETS.keys())
    if 'enabled_os_checks' not in data:
        data['enabled_os_checks'] = list(OS_CHECKS.keys())
    data['auto_refresh'] = data.get('auto_refresh', 0)
    servers.append(data)
    save_servers(servers)
    return jsonify({"status": "ok", "id": data['id']})


@app.route('/api/servers/<server_id>', methods=['PUT'])
@login_required
@permission_required('servers_edit')
def api_update_server(server_id):
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "无效的请求数据"}), 400
    if data.get('persist_enabled') and not load_config().get('log_enabled'):
        return jsonify({"error": "全局日志持久化未启用，请先在系统配置中启用"}), 400
    servers = load_servers()
    for s in servers:
        if s.get('id') == server_id:
            for key in ['name', 'ssh_host', 'ssh_port', 'ssh_user', 'ssh_pass',
                        'db_host', 'db_port', 'db_user', 'db_pass', 'db_name',
                        'isql_cmd', 'enabled_categories', 'enabled_os_checks', 'auto_refresh',
                        'svc_name', 'svc_mgr', 'os_type', 'db_type', 'in_control', 'apps',
                        'svc_start_cmd', 'svc_stop_cmd', 'persist_enabled']:
                if key in data:
                    s[key] = data[key]
            save_servers(servers)
            return jsonify({"status": "ok"})
    return jsonify({"error": "服务不存在"}), 404


@app.route('/api/servers/<server_id>', methods=['DELETE'])
@login_required
@permission_required('servers_edit')
def api_delete_server(server_id):
    servers = load_servers()
    servers = [s for s in servers if s.get('id') != server_id]
    save_servers(servers)
    delete_server_from_db(server_id)
    with CACHE_LOCK:
        CACHE.pop(server_id, None)
    return jsonify({"status": "ok"})


@app.route('/api/servers/<server_id>/collect', methods=['POST'])
@login_required
def api_collect(server_id):
    server = get_server_by_id(server_id)
    if not server:
        return jsonify({"error": "服务不存在"}), 404
    try:
        result = collect_all(server, server.get('enabled_categories'), server.get('enabled_os_checks'))
        with CACHE_LOCK:
            CACHE[server_id] = {"data": result, "time": time.time()}
        _persist_errors(server, result)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500


@app.route('/api/servers/<server_id>/data', methods=['GET'])
@login_required
def api_get_cached(server_id):
    with CACHE_LOCK:
        cached = CACHE.get(server_id)
    if cached:
        return jsonify(cached["data"])
    return jsonify(None)


@app.route('/api/servers/<server_id>/collect-partial', methods=['POST'])
@login_required
def api_collect_partial(server_id):
    server = get_server_by_id(server_id)
    if not server:
        return jsonify({"error": "服务不存在"}), 404
    data = request.get_json(silent=True) or {}
    categories = data.get('categories', [])
    os_checks = data.get('os_checks', [])
    if not categories and not os_checks:
        return jsonify({"error": "请指定采集类别"}), 400
    try:
        result = collect_all(server, categories or [], os_checks or [])
        with CACHE_LOCK:
            existing = CACHE.get(server_id, {"data": {}, "time": 0})
            existing_data = existing.get("data", {})
            if "os_info" in result:
                existing_data["os_info"] = result["os_info"]
            if "db_queries" in result:
                existing_data["db_queries"] = result["db_queries"]
            existing_data["server"] = result["server"]
            existing_data["timestamp"] = result["timestamp"]
            CACHE[server_id] = {"data": existing_data, "time": time.time()}
        _persist_errors(server, result)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500


@app.route('/api/servers/<server_id>/db-control', methods=['POST'])
@login_required
@any_permission_required('control_view', 'control_exec')
def api_db_control(server_id):
    server = get_server_by_id(server_id)
    if not server:
        return jsonify({"error": "服务不存在"}), 404
    data = request.get_json(silent=True) or {}
    action = data.get('action', 'status')
    if action not in ('start', 'stop', 'restart', 'status'):
        return jsonify({"error": "无效的操作"}), 400
    if action != 'status':
        user = session.get('user', {})
        if not has_permission(user, 'control_exec'):
            return jsonify({"error": "无执行权限"}), 403
    try:
        result = db_control(server, action)
        return jsonify(result)
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500


@app.route('/api/servers/<server_id>/app-control', methods=['POST'])
@login_required
@any_permission_required('control_view', 'control_exec')
def api_app_control(server_id):
    server = get_server_by_id(server_id)
    if not server:
        return jsonify({"error": "服务不存在"}), 404
    data = request.get_json(silent=True) or {}
    action = data.get('action', 'status')
    app_name = data.get('app')
    if action not in ('start', 'stop', 'restart', 'status'):
        return jsonify({"error": "无效的操作"}), 400
    if not app_name:
        return jsonify({"error": "请指定应用名称"}), 400
    if action != 'status':
        user = session.get('user', {})
        if not has_permission(user, 'control_exec'):
            return jsonify({"error": "无执行权限"}), 403
    try:
        result = app_control(server, app_name, action)
        return jsonify(result)
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500


@app.route('/config')
@login_required
@permission_required('admin')
def config_page():
    return render_template('config.html', config=load_config(), **template_context())


@app.route('/api/config', methods=['PUT'])
@login_required
@permission_required('admin')
def api_update_config():
    data = request.get_json(silent=True) or {}
    cfg = load_config()
    if 'log_db' in data:
        cfg['log_db'] = data['log_db']
    if 'log_retention_days' in data:
        cfg['log_retention_days'] = data['log_retention_days']
    if 'collect_workers' in data:
        cfg['collect_workers'] = int(data['collect_workers'])
    if 'log_enabled' in data:
        cfg['log_enabled'] = data['log_enabled']
    if 'server_db_enabled' in data:
        cfg['server_db_enabled'] = data['server_db_enabled']
    save_config(cfg)
    _reinit_executor()
    # 使 auth 用户缓存失效（下次请求自动刷新）
    import auth
    auth._users_cache = None
    auth._users_cache_time = 0
    return jsonify({"status": "ok"})


@app.route('/server/<server_id>/log-history')
@login_required
def log_history_page(server_id):
    server = get_server_by_id(server_id)
    if not server:
        return "服务不存在", 404
    return render_template('log_history.html', server=server, **template_context())


@app.route('/api/servers/<server_id>/log-errors', methods=['GET'])
@login_required
def api_log_errors(server_id):
    server = get_server_by_id(server_id)
    if not server:
        return jsonify({"error": "服务不存在"}), 404
    server_name = server.get('name') or server.get('ssh_host', '')
    page = request.args.get('page', 1, type=int)
    size = request.args.get('size', 50, type=int)
    kw = request.args.get('kw', '')
    log_type = request.args.get('type', '')
    all_logs = query_logs(server_name, limit=5000)
    if kw:
        kw_lower = kw.lower()
        all_logs = [l for l in all_logs if kw_lower in (l.get('msg','')+l.get('sql','')+l.get('user','')).lower()]
    if log_type:
        all_logs = [l for l in all_logs if l.get('type') == log_type]
    total = len(all_logs)
    start = (page - 1) * size
    paged = all_logs[start:start + size]
    return jsonify({"logs": paged, "total": total, "server": server_name})


@app.route('/api/config/test-log-db', methods=['POST'])
@login_required
@permission_required('admin')
def api_test_log_db():
    data = request.get_json(silent=True) or {}
    result = {"ssh": None, "db": None}
    ssh_host = data.get('ssh_host', '')

    if ssh_host and ssh_host not in ('127.0.0.1', 'localhost'):
        try:
            client = _ssh_connect({
                'ssh_host': ssh_host,
                'ssh_port': data.get('ssh_port', 22),
                'ssh_user': data.get('ssh_user', 'root'),
                'ssh_pass': data.get('ssh_pass', '')
            })
            client.close()
            result['ssh'] = {"ok": True, "msg": "SSH连接成功"}
        except Exception as e:
            fix_map = SSH_FIX_WIN if ssh_host == '127.0.0.1' else SSH_FIX_LINUX
            result['ssh'] = {"ok": False, "msg": translate_error(str(e), SSH_ERROR_TRANSLATE, fix_map)}
            return jsonify(result)

    try:
        import persist
        persist._exec_sql({
            'host': data.get('host', '127.0.0.1'),
            'port': data.get('port', 2003),
            'user': data.get('user', 'SYSDBA'),
            'pass': data.get('pass', ''),
            'dbname': data.get('dbname', 'OSRDB'),
            'isql': data.get('isql', 'isql'),
            'ssh_host': data.get('ssh_host', ''),
            'ssh_port': data.get('ssh_port', 22),
            'ssh_user': data.get('ssh_user', 'root'),
            'ssh_pass': data.get('ssh_pass', ''),
        }, "select 1;")
        result['db'] = {"ok": True, "msg": "数据库连接成功"}
    except Exception as e:
        result['db'] = {"ok": False, "msg": translate_error(str(e), DB_ERROR_TRANSLATE)}

    return jsonify(result)


@app.route('/api/stream')
@login_required
def api_stream():
    def generate():
        last_seen = {}
        while True:
            servers = load_servers()
            updates = {}
            for s in servers:
                sid = s.get('id')
                with CACHE_LOCK:
                    cached = CACHE.get(sid, {}).get('data')
                if cached:
                    ts = cached.get('timestamp', '')
                    if ts != last_seen.get(sid):
                        last_seen[sid] = ts
                        updates[sid] = cached
            if updates:
                yield f"data: {json.dumps(updates, ensure_ascii=False)}\n\n"
            time.sleep(2)

    return Response(generate(), mimetype='text/event-stream', headers={
        'Cache-Control': 'no-cache',
        'X-Accel-Buffering': 'no',
    })


@app.route('/api/trends/<server_id>')
@login_required
def api_trends(server_id):
    """返回历史趋势数据（CPU/内存/连接数/慢SQL 多指标）"""
    hours = request.args.get('hours', 24, type=int)
    max_points = min(hours * 12, MAX_TREND_POINTS)
    with TREND_LOCK:
        history = TREND_HISTORY.get(server_id, [])
        result = history[-max_points:] if len(history) > max_points else list(history)
    return jsonify(result)


@app.route('/api/servers/<server_id>/health')
@login_required
def api_health_score(server_id):
    """返回服务器健康评分 0-100"""
    with CACHE_LOCK:
        cached = CACHE.get(server_id, {}).get('data')
    if not cached:
        return jsonify({"score": None, "msg": "暂无数据，请先采集"})
    score, details = _calc_health_score(cached)
    return jsonify({"score": score, "details": details})


@app.route('/api/servers/<server_id>/sql-query', methods=['POST'])
@login_required
def api_sql_query(server_id):
    """Web SQL 终端：执行只读查询"""
    server = get_server_by_id(server_id)
    if not server:
        return jsonify({"error": "服务不存在"}), 404
    data = request.get_json(silent=True) or {}
    sql = (data.get('sql', '') or '').strip()
    if not sql:
        return jsonify({"error": "请输入 SQL 语句"}), 400
    sql_upper = sql.upper().strip()
    safe_prefixes = ['SELECT', 'WITH', 'EXPLAIN', 'SHOW', 'DESC', 'DESCRIBE']
    if not any(sql_upper.startswith(p) for p in safe_prefixes):
        return jsonify({"error": "仅支持只读查询 (SELECT / WITH / EXPLAIN / SHOW / DESC)"}), 403
    dangerous = ['DROP', 'DELETE', 'INSERT', 'UPDATE', 'ALTER', 'CREATE', 'TRUNCATE',
                 'GRANT', 'REVOKE', 'EXEC', 'EXECUTE', 'CALL', 'MERGE', 'REPLACE']
    for kw in dangerous:
        if re.search(r'\b' + kw + r'\b', sql_upper):
            return jsonify({"error": f"禁止使用 {kw} 语句"}), 403
    if len(sql) > 5000:
        return jsonify({"error": "SQL 语句过长（最大 5000 字符）"}), 400

    try:
        from collector import _temp_sql_path, _build_sql_cmd, _run_local, _ssh_exec, parse_isql_output
        sql_file = _temp_sql_path(server)
        _, cmd = _build_sql_cmd(server, sql, sql_file)
        if _need_ssh(server):
            client = _ssh_connect(server, timeout=15)
            try:
                out, err, ec = _ssh_exec(client, cmd, timeout=60)
            finally:
                client.close()
        else:
            out, err, ec = _run_local(cmd, timeout=60)
        if ec != 0 and not out.strip():
            return jsonify({"error": err or "执行失败", "raw": out})
        result = parse_isql_output(out, 'sql_query')
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/servers/<server_id>/export')
@login_required
def api_export(server_id):
    """导出采集数据为 CSV"""
    server = get_server_by_id(server_id)
    if not server:
        return jsonify({"error": "服务不存在"}), 404
    with CACHE_LOCK:
        cached = CACHE.get(server_id, {}).get('data')
    if not cached:
        return jsonify({"error": "暂无数据，请先采集"}), 404

    import csv
    import io
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['服务器', server.get('name') or server.get('ssh_host', '')])
    writer.writerow(['采集时间', cached.get('timestamp', '')])
    writer.writerow([])

    # OS info
    os_info = cached.get('os_info', {})
    if os_info:
        writer.writerow(['=== 操作系统信息 ==='])
        for ck, cr in os_info.items():
            writer.writerow([OS_CHECK_LABELS.get(ck, ck)])
            if cr.get('columns') and cr.get('rows'):
                writer.writerow(cr['columns'])
                for row in cr['rows']:
                    writer.writerow(row)
            elif cr.get('output'):
                writer.writerow([cr['output']])
            writer.writerow([])

    # DB queries
    db_queries = cached.get('db_queries', {})
    for cat, queries in db_queries.items():
        writer.writerow([f'=== {QUERY_SETS.get(cat, {}).get("label", cat)} ==='])
        for qn, qr in queries.items():
            writer.writerow([QUERY_LABELS.get(qn, qn)])
            if qr.get('columns') and qr.get('rows'):
                writer.writerow(qr['columns'])
                for row in qr['rows']:
                    writer.writerow(row)
            writer.writerow([])

    csv_content = output.getvalue()
    return Response(
        csv_content,
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename={server_id}_export.csv'}
    )


def auto_collect_job():
    servers = load_servers()
    futures = []
    for server in servers:
        sid = server.get('id')
        with CACHE_LOCK:
            cached = CACHE.get(sid)
        if cached and time.time() - cached.get("time", 0) < 30:
            continue
        futures.append(_get_executor().submit(_collect_one, server))
    for f in as_completed(futures):
        try:
            f.result()
        except Exception:
            pass


def _collect_one(server):
    sid = server.get('id')
    result = collect_all(server, server.get('enabled_categories'), server.get('enabled_os_checks'))
    with CACHE_LOCK:
        CACHE[sid] = {"data": result, "time": time.time()}
    _persist_errors(server, result)
    _record_trend(sid, result)
    return result


def _record_trend(sid, result):
    """从采集结果中提取关键指标，写入趋势历史"""
    try:
        point = {'ts': result.get('timestamp', time.strftime('%Y-%m-%d %H:%M:%S'))}

        # CPU
        cpu_info = result.get('os_info', {}).get('cpu', {})
        cpu_raw = cpu_info.get('output', '') or ''
        cpu_m = re.search(r'LoadPercentage[= ]*(\d+)', cpu_raw)
        if cpu_m:
            point['cpu_pct'] = int(cpu_m.group(1))
        else:
            cpu_m2 = re.search(r'(\d+\.?\d*)\s*id', cpu_raw)
            if cpu_m2:
                point['cpu_pct'] = round(100 - float(cpu_m2.group(1)), 1)

        # Memory
        mem_info = result.get('os_info', {}).get('memory', {})
        mem_raw = mem_info.get('output', '') or ''
        wm = re.search(r'TotalMB=(\d+).*?FreeMB=(\d+).*?UsedMB=(\d+)', mem_raw)
        if wm:
            total_m = int(wm.group(1)); used_m = int(wm.group(3))
            point['mem_pct'] = round(used_m / total_m * 100, 1) if total_m > 0 else 0

        # Sessions
        perf = result.get('db_queries', {}).get('performance', {})
        sessions = perf.get('session_count', {})
        if sessions.get('rows') and sessions['rows']:
            try:
                point['sessions'] = int(sessions['rows'][0][0])
            except (ValueError, IndexError):
                pass

        # Slow SQL count
        slow = perf.get('slow_sql', {})
        if slow.get('rows'):
            point['slow_sql_count'] = len(slow['rows'])

        with TREND_LOCK:
            if sid not in TREND_HISTORY:
                TREND_HISTORY[sid] = []
            TREND_HISTORY[sid].append(point)
            if len(TREND_HISTORY[sid]) > MAX_TREND_POINTS:
                TREND_HISTORY[sid] = TREND_HISTORY[sid][-MAX_TREND_POINTS:]
    except Exception:
        pass


def _calc_health_score(data):
    """计算健康评分 0-100，返回 (score, details)"""
    details = {}
    total_weight = 0
    weighted_score = 0

    # ── CPU (权重 25) ──
    cpu_info = data.get('os_info', {}).get('cpu', {})
    cpu_raw = cpu_info.get('output', '') or ''
    cpu_pct = None
    cm = re.search(r'LoadPercentage[= ]*(\d+)', cpu_raw)
    if cm:
        cpu_pct = int(cm.group(1))
    else:
        cm2 = re.search(r'(\d+\.?\d*)\s*id', cpu_raw)
        if cm2:
            cpu_pct = round(100 - float(cm2.group(1)), 1)
    if cpu_pct is not None:
        cpu_score = max(0, 100 - cpu_pct * 1.2)
        details['cpu'] = {'value': f'{cpu_pct}%', 'score': round(cpu_score)}
        weighted_score += cpu_score * 25
        total_weight += 25

    # ── Memory (权重 25) ──
    mem_info = data.get('os_info', {}).get('memory', {})
    mem_raw = mem_info.get('output', '') or ''
    wm = re.search(r'TotalMB=(\d+).*?FreeMB=(\d+).*?UsedMB=(\d+)', mem_raw)
    if wm:
        total_m = int(wm.group(1)); used_m = int(wm.group(3))
        mem_pct = round(used_m / total_m * 100, 1) if total_m > 0 else 0
        mem_score = max(0, 100 - mem_pct * 1.5)
        details['memory'] = {'value': f'{mem_pct}%', 'score': round(mem_score)}
        weighted_score += mem_score * 25
        total_weight += 25

    # ── Sessions / Connections (权重 20) ──
    perf = data.get('db_queries', {}).get('performance', {})
    sessions = perf.get('session_count', {})
    if sessions.get('rows') and sessions['rows']:
        try:
            sess_count = int(sessions['rows'][0][0])
            sess_score = max(0, 100 - sess_count * 0.5)  # 200连接=0分
            details['sessions'] = {'value': str(sess_count), 'score': round(sess_score)}
            weighted_score += sess_score * 20
            total_weight += 20
        except (ValueError, IndexError):
            pass

    # ── Slow SQL (权重 15) ──
    slow = perf.get('slow_sql', {})
    slow_count = len(slow.get('rows', [])) if slow.get('rows') else 0
    slow_score = max(0, 100 - slow_count * 10)
    details['slow_sql'] = {'value': f'{slow_count} 条', 'score': round(slow_score)}
    weighted_score += slow_score * 15
    total_weight += 15

    # ── Deadlocks (权重 15) ──
    deadlock = perf.get('deadlock_count', {})
    dl_count = 0
    if deadlock.get('rows') and deadlock['rows']:
        try:
            dl_count = int(deadlock['rows'][0][0])
        except (ValueError, IndexError):
            pass
    dl_score = 100 if dl_count == 0 else max(0, 100 - dl_count * 20)
    details['deadlocks'] = {'value': str(dl_count), 'score': round(dl_score)}
    weighted_score += dl_score * 15
    total_weight += 15

    if total_weight == 0:
        return 100, details
    score = round(weighted_score / total_weight)
    return max(0, min(100, score)), details


def _persist_errors(server, result):
    if not server.get('persist_enabled'):
        return
    _get_persist_executor().submit(_do_persist, server, result)


def _do_persist(server, result):
    server_name = server.get('name') or server.get('ssh_host', '')

    if result.get('os_info'):
        for ck, cr in result['os_info'].items():
            text = cr.get('output', '')
            if not text:
                continue
            if ck in ('os_errors',):
                for line in text.split('\n'):
                    line = line.strip()
                    if line and not line.startswith('==='):
                        persist_os_error(server_name, ck, line)
            elif ck == 'db_log_errors':
                _parse_elog_content(server_name, text)

    perf = result.get('db_queries', {}).get('performance', {})
    slow = perf.get('slow_sql', {})
    if slow and slow.get('rows'):
        for row in slow['rows']:
            if row and len(row) >= 2:
                cost = float(row[0]) if row[0] else 0
                sql = str(row[1] or '')[:800]
                if cost >= 0.5 and sql:
                    persist_slow_sql(server_name, {'cost': cost, 'sql': sql})


def _parse_elog_content(server_name, text):
    blocks = text.split('===FILE:')
    for block in blocks:
        if not block.strip():
            continue
        lines = block.split('\n')
        filepath = lines[0].replace('===', '').strip() if lines else ''
        content_lines = lines[1:]
        entry = {}
        for line in content_lines:
            line = line.strip()
            if not line:
                if entry and entry.get('msg'):
                    entry['tool'] = filepath.split('/')[-1][:100] if filepath else ''
                    persist_db_error(server_name, entry)
                    entry = {}
                continue
            msg_match = re.match(r'^(ERROR|FATAL|WARNING|PANIC)\s*[:\-]?\s*(.+)', line, re.IGNORECASE)
            if msg_match:
                if entry and entry.get('msg'):
                    entry['tool'] = filepath.split('/')[-1][:100] if filepath else ''
                    persist_db_error(server_name, entry)
                entry = {'msg': msg_match.group(2)[:500]}
                continue
            t_match = re.search(r'(\d{4}[-/]\d{2}[-/]\d{2}\s+\d{2}:\d{2}:\d{2})', line)
            if t_match and not entry.get('time'):
                entry['time'] = t_match.group(1)
            u_match = re.search(r'(?:user|USER|User)\s*[=:]\s*(\w+)', line)
            if u_match and not entry.get('user'):
                entry['user'] = u_match.group(1)
            cost_match = re.search(r'(?:duration|cost|elapsed|耗时)\s*[=:]\s*([\d.]+)\s*(ms|s|秒|毫秒)?', line, re.IGNORECASE)
            if cost_match:
                cost = float(cost_match.group(1))
                if cost_match.group(2) in ('ms', '毫秒'):
                    cost = cost / 1000
                entry['cost'] = round(cost, 3)
            sql_match = re.search(r'(?:statement|sql|query|SQL|语句)\s*[=:]\s*(.+)', line, re.IGNORECASE)
            if sql_match and not entry.get('sql'):
                entry['sql'] = sql_match.group(1)[:800]
            if not entry.get('msg'):
                entry['msg'] = line[:500]
        if entry and entry.get('msg'):
            entry['tool'] = filepath.split('/')[-1][:100] if filepath else ''
            persist_db_error(server_name, entry)


if scheduler is not None:
    scheduler.add_job(auto_collect_job, 'interval', seconds=30)
    scheduler.add_job(cleanup_old_logs, 'interval', hours=24)


# ═══════════════════════════════════════════════
#  巡检导出中心
# ═══════════════════════════════════════════════
REPORTS_DIR = os.path.join(DATA_DIR, 'reports')
EXPORT_CATEGORIES = [
    {'id': 'basic_info', 'label': '基础信息', 'icon': 'bi-info-circle'},
    {'id': 'db_info', 'label': '数据库信息', 'icon': 'bi-database'},
    {'id': 'storage', 'label': '存储空间', 'icon': 'bi-hdd'},
    {'id': 'objects', 'label': '对象统计', 'icon': 'bi-grid'},
    {'id': 'performance', 'label': '性能监控', 'icon': 'bi-graph-up'},
    {'id': 'install_path', 'label': '安装路径', 'icon': 'bi-folder'},
    {'id': 'db_log_errors', 'label': '数据库日志', 'icon': 'bi-file-earmark-text'},
]
EXPORT_OS_CHECKS = [
    {'id': 'memory', 'label': '系统内存', 'icon': 'bi-memory'},
    {'id': 'disk', 'label': '系统磁盘', 'icon': 'bi-hdd-stack'},
    {'id': 'cpu', 'label': 'CPU负载', 'icon': 'bi-cpu'},
    {'id': 'os_errors', 'label': '系统日志错误', 'icon': 'bi-exclamation-diamond'},
]

# ═══ 应用模板 ═══
APP_TEMPLATES_FILE = os.path.join(DATA_DIR, 'app_templates.json')

def load_app_templates():
    if os.path.exists(APP_TEMPLATES_FILE):
        with open(APP_TEMPLATES_FILE, 'r', encoding='utf-8') as f:
            return json.load(f).get('templates', {})
    return {}

@app.route('/api/app-templates')
@login_required
def api_app_templates():
    return jsonify(load_app_templates())


@app.route('/reports')
@login_required
def reports_page():
    servers = load_servers()
    cfg = load_config()
    schedule_cfg = cfg.get('export_schedule', {})
    return render_template('reports.html',
                           servers=servers,
                           categories=EXPORT_CATEGORIES,
                           os_checks=EXPORT_OS_CHECKS,
                           schedule=schedule_cfg,
                           **template_context())


@app.route('/api/export/sql-result', methods=['POST'])
@login_required
def api_export_sql_result():
    """将 SQL 查询结果导出为 Excel"""
    data = request.get_json(silent=True) or {}
    columns = data.get('columns', [])
    rows = data.get('rows', [])
    if not columns:
        return jsonify({"error": "无列信息"}), 400

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "SQL查询结果"
        # 表头
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
        for ci, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=ci, value=str(col))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        # 数据行
        for ri, row in enumerate(rows, 2):
            for ci in range(len(columns)):
                val = row[ci] if ci < len(row) else ''
                ws.cell(row=ri, column=ci + 1, value=str(val) if val is not None else 'NULL')
        # 调整列宽
        for ci in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 18

        report_dir = os.path.join(DATA_DIR, 'reports')
        os.makedirs(report_dir, exist_ok=True)
        filename = f"sql_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(report_dir, filename)
        wb.save(filepath)
        return jsonify({"status": "ok", "filename": filename})
    except Exception as e:
        return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500


@app.route('/api/export/xlsx', methods=['POST'])
@login_required
def api_export_xlsx():
    data = request.get_json(silent=True) or {}
    server_ids = data.get('servers', [])
    categories = data.get('categories', [])
    os_checks = data.get('os_checks', [])
    organize_by = data.get('organize_by', 'server')  # 'server' or 'category'

    if not server_ids:
        return jsonify({"error": "请选择至少一台服务器"}), 400
    if not categories and not os_checks:
        return jsonify({"error": "请选择至少一项导出内容"}), 400

    servers = [s for s in load_servers() if s.get('id') in server_ids]
    if not servers:
        return jsonify({"error": "未找到所选服务器"}), 404

    try:
        wb = Workbook()
        wb.remove(wb.active)

        header_font = Font(name='Microsoft YaHei', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
        section_font = Font(name='Microsoft YaHei', bold=True, size=12, color='4F46E5')
        thin_border = Border(
            left=Side(style='thin', color='D4D4D8'),
            right=Side(style='thin', color='D4D4D8'),
            top=Side(style='thin', color='D4D4D8'),
            bottom=Side(style='thin', color='D4D4D8'),
        )
        wrap_align = Alignment(wrap_text=True, vertical='top')

        if organize_by == 'server':
            for server in servers:
                _fill_server_sheet(wb, server, categories, os_checks,
                                   header_font, header_fill, section_font, thin_border, wrap_align)
            # 汇总 Sheet
            _fill_summary_sheet(wb, servers, categories, os_checks,
                                header_font, header_fill, section_font, thin_border)
        else:
            # 按类别组织
            for cat in categories:
                if cat in QUERY_SETS:
                    ws = wb.create_sheet(title=QUERY_SETS[cat]['label'][:31])
                    _fill_category_sheet(ws, servers, cat, 'db',
                                         header_font, header_fill, section_font, thin_border, wrap_align)
            for chk in os_checks:
                if chk in OS_CHECK_LABELS:
                    ws = wb.create_sheet(title=OS_CHECK_LABELS[chk][:31])
                    _fill_category_sheet(ws, servers, chk, 'os',
                                         header_font, header_fill, section_font, thin_border, wrap_align)

        # 保存文件
        os.makedirs(REPORTS_DIR, exist_ok=True)
        ts = time.strftime('%Y%m%d_%H%M%S')
        filename = f'巡检报告_{ts}.xlsx'
        filepath = os.path.join(REPORTS_DIR, filename)
        wb.save(filepath)

        # 记录历史
        _add_export_history(filename, len(servers), len(categories) + len(os_checks))

        resp = send_file(filepath, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        from urllib.parse import quote
        resp.headers['X-Filename'] = quote(filename, safe='')
        return resp

    except Exception as e:
        return jsonify({"error": f"导出失败: {str(e)}"}), 500


def _fill_server_sheet(wb, server, categories, os_checks, hf, hfill, sf, border, align):
    name = server.get('name') or server.get('ssh_host', '')
    sheet_name = (name[:28] + '..') if len(name) > 31 else name
    sheet_name = sheet_name[:31]
    ws = wb.create_sheet(title=sheet_name)

    sid = server.get('id')
    with CACHE_LOCK:
        cached = CACHE.get(sid, {}).get('data')

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell = ws.cell(row=row, column=1, value=f'📊 巡检报告 — {name}')
    cell.font = Font(name='Microsoft YaHei', bold=True, size=14, color='4F46E5')
    row += 1
    ws.cell(row=row, column=1, value=f'采集时间: {cached.get("timestamp", "暂无数据") if cached else "暂无数据"}').font = Font(color='64748B', size=10)
    row += 2

    if not cached:
        ws.cell(row=row, column=1, value='暂无采集数据，请先执行采集').font = Font(color='DC2626', size=11)
        _auto_width(ws)
        return

    os_info = cached.get('os_info', {})
    db_queries = cached.get('db_queries', {})

    for chk in os_checks:
        if chk in os_info and chk in OS_CHECK_LABELS:
            row = _write_section_header(ws, row, f'🖥 {OS_CHECK_LABELS[chk]}', sf)
            row = _write_check_result(ws, row, os_info[chk], hf, hfill, border, align)
            row += 1

    for cat in categories:
        # 先查 db_queries（数据库采集），再查 os_info（系统采集但归入数据库分类的项）
        if cat in db_queries and cat in QUERY_SETS:
            row = _write_section_header(ws, row, f'📋 {QUERY_SETS[cat]["label"]}', sf)
            for qname, qresult in db_queries[cat].items():
                label = QUERY_LABELS.get(qname, qname)
                row = _write_section_header(ws, row, label, Font(name='Microsoft YaHei', bold=True, size=10, color='6366F1'))
                row = _write_query_result(ws, row, qresult, hf, hfill, border, align)
                row += 1
        elif cat in os_info and cat in OS_CHECK_LABELS:
            row = _write_section_header(ws, row, f'📋 {OS_CHECK_LABELS[cat]}', sf)
            row = _write_check_result(ws, row, os_info[cat], hf, hfill, border, align)
            row += 1

    _auto_width(ws)


def _fill_category_sheet(ws, servers, key, stype, hf, hfill, sf, border, align):
    row = 1
    label = QUERY_SETS[key]['label'] if stype == 'db' else OS_CHECK_LABELS.get(key, key)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1, value=f'📊 {label}').font = Font(name='Microsoft YaHei', bold=True, size=14, color='4F46E5')
    row += 2

    for server in servers:
        name = server.get('name') or server.get('ssh_host', '')
        sid = server.get('id')
        with CACHE_LOCK:
            cached = CACHE.get(sid, {}).get('data')
        row = _write_section_header(ws, row, f'🖥 {name}', sf)
        if not cached:
            ws.cell(row=row, column=1, value='暂无数据').font = Font(color='DC2626', size=10)
            row += 2
            continue

        if stype == 'os':
            chk_data = cached.get('os_info', {}).get(key)
            if chk_data:
                row = _write_check_result(ws, row, chk_data, hf, hfill, border, align)
            else:
                ws.cell(row=row, column=1, value='(未采集)').font = Font(color='94A3B8', size=10)
                row += 1
        else:
            # 数据库类别：先查 db_queries，再查 os_info
            if key in cached.get('db_queries', {}):
                cat_data = cached['db_queries'][key]
                for qname, qresult in cat_data.items():
                    qlabel = QUERY_LABELS.get(qname, qname)
                    ws.cell(row=row, column=1, value=qlabel).font = Font(name='Microsoft YaHei', bold=True, size=10, color='6366F1')
                    row += 1
                    row = _write_query_result(ws, row, qresult, hf, hfill, border, align)
            elif key in cached.get('os_info', {}):
                chk_data = cached['os_info'][key]
                row = _write_check_result(ws, row, chk_data, hf, hfill, border, align)
            else:
                ws.cell(row=row, column=1, value='(未采集)').font = Font(color='94A3B8', size=10)
                row += 1
        row += 1
    _auto_width(ws)


def _fill_summary_sheet(wb, servers, categories, os_checks, hf, hfill, sf, border):
    ws = wb.create_sheet(title='汇总概览')
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1, value='📊 巡检汇总报告').font = Font(name='Microsoft YaHei', bold=True, size=14, color='4F46E5')
    row += 1
    ws.cell(row=row, column=1, value=f'导出时间: {time.strftime("%Y-%m-%d %H:%M:%S")}').font = Font(color='64748B', size=10)
    row += 2

    # 服务器总览
    headers = ['服务器', '采集时间', '在线状态', '连接数', '死锁数', '慢SQL数']
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = hf; cell.fill = hfill; cell.border = border
    row += 1

    for server in servers:
        name = server.get('name') or server.get('ssh_host', '')
        sid = server.get('id')
        with CACHE_LOCK:
            cached = CACHE.get(sid, {}).get('data')
        ts = cached.get('timestamp', '-') if cached else '-'

        sessions = '-'
        deadlocks = '-'
        slow_count = '-'
        if cached:
            perf = cached.get('db_queries', {}).get('performance', {})
            sc = perf.get('session_count', {}).get('rows')
            if sc and sc[0]:
                sessions = str(sc[0][0])
            dl = perf.get('deadlock_count', {}).get('rows')
            if dl and dl[0]:
                deadlocks = str(dl[0][0])
            sl = perf.get('slow_sql', {}).get('rows', [])
            slow_count = str(len(sl))

        vals = [name, ts, '在线' if cached else '未采集', sessions, deadlocks, slow_count]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.border = border; cell.alignment = Alignment(vertical='top')
        row += 1

    row += 1
    ws.cell(row=row, column=1, value=f'导出内容: 数据库类别 {len(categories)} 项, 系统检查 {len(os_checks)} 项').font = Font(color='94A3B8', size=9)
    _auto_width(ws)


def _write_section_header(ws, row, text, font):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1, value=text).font = font
    return row + 1


def _write_query_result(ws, row, qr, hf, hfill, border, align):
    if not qr or not qr.get('columns'):
        ws.cell(row=row, column=1, value='(无数据)').font = Font(color='94A3B8', size=10)
        return row + 1
    for ci, col in enumerate(qr['columns'], 1):
        cell = ws.cell(row=row, column=ci, value=str(col))
        cell.font = hf; cell.fill = hfill; cell.border = border
    row += 1
    for r_data in (qr.get('rows') or []):
        for ci, val in enumerate(r_data, 1):
            cell = ws.cell(row=row, column=ci, value=str(val) if val is not None else '')
            cell.border = border; cell.alignment = align
        row += 1
    return row


def _write_check_result(ws, row, cr, hf, hfill, border, align):
    if cr.get('columns') and cr.get('rows'):
        return _write_query_result(ws, row, cr, hf, hfill, border, align)
    output = cr.get('output', '')
    if output:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        cell = ws.cell(row=row, column=1, value=output[:3000])
        cell.font = Font(name='Consolas', size=9, color='334155')
        cell.alignment = Alignment(wrap_text=True)
        return row + 1
    ws.cell(row=row, column=1, value='(无数据)').font = Font(color='94A3B8', size=10)
    return row + 1


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def _add_export_history(filename, server_count, item_count):
    history_file = os.path.join(REPORTS_DIR, 'history.json')
    history = []
    if os.path.exists(history_file):
        try:
            with open(history_file, 'r', encoding='utf-8') as f:
                history = json.load(f)
        except Exception:
            history = []
    history.insert(0, {
        'filename': filename,
        'server_count': server_count,
        'item_count': item_count,
        'time': time.strftime('%Y-%m-%d %H:%M:%S'),
        'size': os.path.getsize(os.path.join(REPORTS_DIR, filename)),
    })
    # 最多保留 100 条历史
    if len(history) > 100:
        old = history.pop()
        old_path = os.path.join(REPORTS_DIR, old['filename'])
        if os.path.exists(old_path):
            os.remove(old_path)
    with open(history_file, 'w', encoding='utf-8') as f:
        json.dump(history, f, ensure_ascii=False, indent=2)


@app.route('/api/export/history', methods=['GET'])
@login_required
def api_export_history():
    history_file = os.path.join(REPORTS_DIR, 'history.json')
    if not os.path.exists(history_file):
        return jsonify([])
    with open(history_file, 'r', encoding='utf-8') as f:
        history = json.load(f)
    for h in history:
        h['size_str'] = _format_size(h.get('size', 0))
    return jsonify(history)


@app.route('/api/export/history/<filename>', methods=['DELETE'])
@login_required
def api_delete_export_history(filename):
    filepath = os.path.join(REPORTS_DIR, filename)
    if os.path.exists(filepath):
        os.remove(filepath)
    history_file = os.path.join(REPORTS_DIR, 'history.json')
    if os.path.exists(history_file):
        with open(history_file, 'r', encoding='utf-8') as f:
            history = json.load(f)
        history = [h for h in history if h.get('filename') != filename]
        with open(history_file, 'w', encoding='utf-8') as f:
            json.dump(history, f, ensure_ascii=False, indent=2)
    return jsonify({"status": "ok"})


@app.route('/api/export/xlsx-download/<path:filename>')
@app.route('/api/export/download/<path:filename>')
@login_required
def api_download_export(filename):
    filepath = os.path.join(REPORTS_DIR, filename)
    if not os.path.exists(filepath):
        return jsonify({"error": "文件不存在"}), 404
    return send_file(filepath, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/export/schedule', methods=['GET', 'PUT'])
@login_required
def api_export_schedule():
    if request.method == 'GET':
        cfg = load_config()
        return jsonify(cfg.get('export_schedule', {}))

    data = request.get_json(silent=True) or {}
    cfg = load_config()
    cfg['export_schedule'] = {
        'enabled': data.get('enabled', False),
        'frequency': data.get('frequency', 'daily'),  # daily / weekly / monthly
        'time': data.get('time', '08:00'),
        'keep': data.get('keep', 30),
    }
    save_config(cfg)

    # 更新定时任务
    if scheduler is not None:
        for job in scheduler.get_jobs():
            if job.id == 'auto_export':
                job.remove()
        if cfg['export_schedule'].get('enabled'):
            freq = cfg['export_schedule'].get('frequency', 'daily')
            time_str = cfg['export_schedule'].get('time', '08:00')
            hour, minute = map(int, time_str.split(':'))
            if freq == 'daily':
                scheduler.add_job(_auto_export_job, 'cron', hour=hour, minute=minute, id='auto_export')
            elif freq == 'weekly':
                scheduler.add_job(_auto_export_job, 'cron', day_of_week='mon', hour=hour, minute=minute, id='auto_export')
            elif freq == 'monthly':
                scheduler.add_job(_auto_export_job, 'cron', day=1, hour=hour, minute=minute, id='auto_export')

    return jsonify({"status": "ok"})


def _auto_export_job():
    """定时导出任务：导出所有服务器的全部数据"""
    servers = load_servers()
    if not servers:
        return
    try:
        wb = Workbook()
        wb.remove(wb.active)
        hf = Font(name='Microsoft YaHei', bold=True, size=11, color='FFFFFF')
        hfill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
        sf = Font(name='Microsoft YaHei', bold=True, size=12, color='4F46E5')
        border = Border(
            left=Side(style='thin', color='D4D4D8'),
            right=Side(style='thin', color='D4D4D8'),
            top=Side(style='thin', color='D4D4D8'),
            bottom=Side(style='thin', color='D4D4D8'),
        )
        align = Alignment(wrap_text=True, vertical='top')
        all_cats = list(QUERY_SETS.keys())
        all_os = list(OS_CHECKS.keys())

        for server in servers:
            _fill_server_sheet(wb, server, all_cats, all_os, hf, hfill, sf, border, align)
        _fill_summary_sheet(wb, servers, all_cats, all_os, hf, hfill, sf, border)

        os.makedirs(REPORTS_DIR, exist_ok=True)
        ts = time.strftime('%Y%m%d_%H%M%S')
        filename = f'巡检报告_自动_{ts}.xlsx'
        filepath = os.path.join(REPORTS_DIR, filename)
        wb.save(filepath)
        _add_export_history(filename, len(servers), len(all_cats) + len(all_os))

        # 清理旧报告
        cfg = load_config()
        keep = cfg.get('export_schedule', {}).get('keep', 30)
        _cleanup_old_reports(keep)
    except Exception:
        pass


def _cleanup_old_reports(keep):
    history_file = os.path.join(REPORTS_DIR, 'history.json')
    if not os.path.exists(history_file):
        return
    with open(history_file, 'r', encoding='utf-8') as f:
        history = json.load(f)
    while len(history) > keep:
        old = history.pop()
        old_path = os.path.join(REPORTS_DIR, old['filename'])
        if os.path.exists(old_path):
            os.remove(old_path)
    with open(history_file, 'w', encoding='utf-8') as f:
        json.dump(history, f, ensure_ascii=False, indent=2)


def _format_size(size_bytes):
    if size_bytes < 1024:
        return f'{size_bytes} B'
    elif size_bytes < 1024 * 1024:
        return f'{size_bytes / 1024:.1f} KB'
    else:
        return f'{size_bytes / (1024 * 1024):.1f} MB'


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='管控平台')
    parser.add_argument('--port', type=int, default=5080, help='监听端口 (默认: 5080)')
    args = parser.parse_args()

    if getattr(sys, 'frozen', False):
        import webbrowser
        threading.Thread(target=lambda: (time.sleep(1.5), webbrowser.open(f'http://localhost:{args.port}')), daemon=True).start()

    app.run(host='0.0.0.0', port=args.port, debug=False, use_reloader=False)
